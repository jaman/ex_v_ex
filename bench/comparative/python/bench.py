#!/usr/bin/env python3
"""Comparative xlsx benchmark harness - Python / openpyxl implementation.

CLI: bench.py <spec.json> <mode:create|edit> <input|-> <output>
Emits one JSON line to stdout with wall time + cell counts.
"""
from __future__ import annotations

import json
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook


def gen_cell(row: int, col: int):
    if col == 1:
        return f"SYM{row:04d}"
    if col == 2:
        return row * 1.5 + 2 * 0.25
    if col == 3:
        return row * 0.1 + 3 * 0.01
    if col == 4:
        return row * 4
    if col == 5:
        return f"LBL-{row:04d}"
    if col == 6:
        return row % 2 == 0
    if col == 7:
        return "B"
    return row + col


def write_data_sheet(ws, rows: int, cols: int) -> int:
    count = 0
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            ws.cell(row=row, column=col, value=gen_cell(row, col))
            count += 1
    return count


def clear_data_sheet(ws, rows: int, cols: int) -> int:
    count = 0
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            ws.cell(row=row, column=col, value=None)
            count += 1
    return count


def create(spec: dict, output: Path) -> tuple[int, int]:
    wb = Workbook()
    wb.remove(wb.active)

    written = 0

    for s in spec["template"]["summary_sheets"]:
        wb.create_sheet(s["name"])

    for s in spec["template"]["data_sheets"]:
        ws = wb.create_sheet(s["name"])
        written += write_data_sheet(ws, s["rows"], s["cols"])

    for s in spec["template"]["reference_sheets"]:
        ws = wb.create_sheet(s["name"])
        written += write_data_sheet(ws, s["rows"], s["cols"])

    for s in spec["template"]["summary_sheets"]:
        ws = wb[s["name"]]
        for rng in s.get("merges", []) or []:
            ws.merge_cells(rng)
        for cell, formula in s.get("formulas", []) or []:
            ws[cell] = formula

    wb.save(output)
    return written, 0


def edit(spec: dict, input_path: Path, output: Path) -> tuple[int, int]:
    wb = load_workbook(input_path)
    rows = spec["manipulation"]["rows_per_data_sheet"]
    cols = spec["manipulation"]["cols_per_data_sheet"]

    cleared = 0
    written = 0

    data_names = [s["name"] for s in spec["template"]["data_sheets"]]
    present = [n for n in data_names if n in wb.sheetnames]

    for name in present:
        cleared += clear_data_sheet(wb[name], rows, cols)

    for name in present:
        written += write_data_sheet(wb[name], rows, cols)

    wb.save(output)
    return written, cleared


def main() -> int:
    spec_path, mode, input_, output = sys.argv[1:5]
    spec = json.loads(Path(spec_path).read_text())

    start = time.perf_counter()
    if mode == "create":
        written, cleared = create(spec, Path(output))
    elif mode == "edit":
        written, cleared = edit(spec, Path(input_), Path(output))
    else:
        print(f"unknown mode: {mode}", file=sys.stderr)
        return 2
    elapsed_ms = round((time.perf_counter() - start) * 1000, 3)

    print(
        json.dumps(
            {
                "lang": "python",
                "mode": mode,
                "wall_ms": elapsed_ms,
                "cells_written": written,
                "cells_cleared": cleared,
            }
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
